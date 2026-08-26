"""Tests for report assembly, exports and the example dataset."""

from __future__ import annotations

import json

import pytest

from calculations import analyze
from example_data import inputs_from_dict, inputs_to_dict, load_examples
from explanations import EXPLANATIONS, explanations_frame
from export import build_report, bundle_to_csv, bundle_to_excel, bundle_to_pdf
from risk import compute_risk_metrics, evaluate_decision
from scenarios import (
    MonteCarloConfig,
    run_monte_carlo,
    run_scenarios,
    sensitivity_gmp_vs_probability,
)
from tests.test_calculations import frictionless
from validation import validate_inputs


@pytest.fixture(scope="module")
def bundle():
    inputs = frictionless(gmp=25.0, probability=0.2, n_accounts=2)
    result = analyze(inputs)
    risk = compute_risk_metrics(result)
    decision = evaluate_decision(result, risk)
    return build_report(
        result,
        risk,
        decision,
        scenarios=run_scenarios(inputs),
        sensitivities={
            "GMP vs probability": sensitivity_gmp_vs_probability(
                inputs, [0.0, 25.0, 50.0], [0.05, 0.1, 0.2]
            )
        },
        monte_carlo=run_monte_carlo(inputs, MonteCarloConfig(n_simulations=2_000)),
    )


def test_report_contains_every_required_section(bundle):
    for section in (
        "Summary",
        "Decision",
        "Decision checks",
        "Assumptions",
        "Accounts",
        "Capital & financing",
        "Risk metrics",
        "Scenarios",
        "Monte Carlo",
        "Method notes",
    ):
        assert section in bundle.tables
    assert any(name.startswith("Sensitivity") for name in bundle.tables)


def test_csv_export_is_readable_and_labelled(bundle):
    text = bundle_to_csv(bundle)
    assert text.startswith("# IPO financing analysis")
    assert "## Summary" in text
    assert "## Risk metrics" in text
    assert text.count("##") >= len(bundle.tables)


def test_excel_export_produces_a_valid_workbook(bundle):
    payload = bundle_to_excel(bundle)
    assert payload[:2] == b"PK"  # xlsx is a zip archive
    from openpyxl import load_workbook
    import io

    workbook = load_workbook(io.BytesIO(payload))
    assert "Summary" in workbook.sheetnames
    assert all(len(name) <= 31 for name in workbook.sheetnames)
    assert len(set(workbook.sheetnames)) == len(workbook.sheetnames)


def test_pdf_export_has_a_valid_structure(bundle):
    payload = bundle_to_pdf(bundle)
    assert payload.startswith(b"%PDF-1.4")
    assert payload.rstrip().endswith(b"%%EOF")
    assert b"/Type /Catalog" in payload
    assert b"/Type /Page " in payload

    # Every cross-reference offset must point at its object header.
    start = payload.rindex(b"startxref")
    xref_offset = int(payload[start + len("startxref"):].split()[0])
    assert payload[xref_offset : xref_offset + 4] == b"xref"
    lines = payload[xref_offset:].split(b"\n")
    count = int(lines[1].split()[1])
    for index, line in enumerate(lines[2 : 2 + count]):
        if line.endswith(b"f "):
            continue
        offset = int(line.split()[0])
        assert payload[offset : offset + len(f"{index} 0 obj")] == f"{index} 0 obj".encode()


def test_pdf_escapes_characters_that_would_corrupt_the_file():
    from export import _pdf_escape

    assert _pdf_escape("a (b) \\c") == r"a \(b\) \\c"


def test_pdf_can_be_limited_to_selected_tables(bundle):
    payload = bundle_to_pdf(bundle, include_tables=["Summary"])
    assert len(payload) < len(bundle_to_pdf(bundle))


# ---------------------------------------------------------------------------
# Example dataset
# ---------------------------------------------------------------------------
def test_every_bundled_example_is_valid_and_analysable():
    examples = load_examples()
    assert len(examples) >= 3
    for example in examples:
        report = validate_inputs(example.inputs)
        assert report.is_valid, f"{example.name}: {[str(i) for i in report.errors]}"
        result = analyze(example.inputs)
        assert result.capital.total_application_amount > 0
        assert example.notes


def test_the_examples_span_the_decision_range():
    verdicts = set()
    for example in load_examples():
        result = analyze(example.inputs)
        risk = compute_risk_metrics(result)
        verdicts.add(evaluate_decision(result, risk).verdict.value)
    assert "GO" in verdicts
    assert "NO-GO" in verdicts


def test_inputs_survive_a_serialisation_round_trip():
    original = load_examples()[1].inputs
    restored = inputs_from_dict(json.loads(json.dumps(inputs_to_dict(original))))
    assert restored == original
    assert analyze(restored).expected_net_profit == pytest.approx(
        analyze(original).expected_net_profit
    )


# ---------------------------------------------------------------------------
# Method documentation
# ---------------------------------------------------------------------------
def test_every_metric_explanation_is_complete():
    for entry in EXPLANATIONS:
        assert entry.formula.strip()
        assert entry.inputs_used.strip()
        assert entry.interpretation.strip()
        assert entry.limitations.strip()


def test_explanations_cover_the_headline_metrics():
    metrics = {entry.metric for entry in EXPLANATIONS}
    assert any("Expected profit" in m for m in metrics)
    assert any("Break-even" in m for m in metrics)
    assert any("Financing" in m for m in metrics)
    assert any("Monte Carlo" in m for m in metrics)
    assert explanations_frame().shape[0] == len(EXPLANATIONS)


def test_expected_value_is_explained_as_not_being_a_guarantee():
    entry = next(e for e in EXPLANATIONS if e.metric == "Expected profit")
    assert "not the outcome of any single application" in entry.interpretation
